"""Import / export helpers for Contacts.

This module is intentionally dependency-light:
  * CSV uses :mod:`csv` (stdlib)
  * JSON uses :mod:`json` (stdlib)
  * XLSX uses :mod:`openpyxl` (third-party — added to requirements)
  * vCard is hand-rolled text parsing; vCard 3.0 is line-based and we don't
    need a full RFC parser for the properties we support.

The shared public surface is:

    export_contacts(contacts, fmt)        -> HttpResponse
    import_contacts(file_obj, fmt, ws, u) -> dict  (summary)
    template_for(fmt)                     -> HttpResponse

Column mapping for CSV / XLSX is driven by ``COLUMN_ALIASES``: every accepted
header (English + Spanish + common abbreviations) maps to a canonical field
on the :class:`crm.models.Contact` model.  Order does not matter — the
importer picks columns by header name, so the user can rearrange or omit
columns in their spreadsheet and it will still work.

vCard is a different beast: it has its own key standard (FN, N, EMAIL, TEL,
ORG, ADR, BDAY, URL, NOTE, X-...).  We only follow the standard properties;
unknown X- properties are stored in the contact's
``custom_social_profiles`` JSON.
"""

from __future__ import annotations

import csv
import io
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Iterable, Mapping

from django.http import HttpResponse

from .models import City, Company, Contact, Country, State
from .utils import get_or_create_locale


# ---------------------------------------------------------------------------
# Canonical column order (used for templates and exports)
# ---------------------------------------------------------------------------
CONTACT_FIELDS: list[tuple[str, str]] = [
    ("first_name", "First Name"),
    ("middle_name", "Middle Name"),
    ("last_name", "Last Name"),
    ("nickname", "Nickname"),
    ("gender", "Gender"),
    ("birthday", "Birthday"),
    ("personal_email", "Personal Email"),
    ("work_email", "Work Email"),
    ("personal_phone", "Personal Phone"),
    ("work_phone", "Work Phone"),
    ("company", "Company"),
    ("website", "Website"),
    ("address", "Address"),
    ("country", "Country"),
    ("state", "State / Province"),
    ("city", "City"),
    ("notes", "Notes"),
    ("linkedin", "LinkedIn"),
    ("github", "GitHub"),
    ("facebook", "Facebook"),
    ("instagram", "Instagram"),
    ("twitter", "Twitter / X"),
    ("tiktok", "TikTok"),
    ("youtube", "YouTube"),
    ("telegram", "Telegram"),
    ("discord", "Discord"),
]

FIELD_DESCRIPTIONS: dict[str, str] = {
    "first_name": "Required. Given / first name.",
    "middle_name": "Optional. Middle name.",
    "last_name": "Optional. Family / last name.",
    "nickname": "Optional. Informal name.",
    "gender": "Optional. MALE / FEMALE / OTHER / PREFER_NOT_TO_SAY.",
    "birthday": "Optional. YYYY-MM-DD.",
    "personal_email": "Optional. Personal email address.",
    "work_email": "Optional. Work / business email address.",
    "personal_phone": "Optional. Mobile / cell phone.",
    "work_phone": "Optional. Office phone or extension.",
    "company": "Optional. Company name. Created if it does not exist.",
    "website": "Optional. Personal website URL.",
    "address": "Optional. Postal address (free text).",
    "country": "Optional. Country name (Spanish or English).",
    "state": "Optional. State / province name.",
    "city": "Optional. City name.",
    "notes": "Optional. Free-form notes.",
    "linkedin": "Optional. Full LinkedIn profile URL.",
    "github": "Optional. Full GitHub profile URL.",
    "facebook": "Optional. Full Facebook profile URL.",
    "instagram": "Optional. Full Instagram profile URL.",
    "twitter": "Optional. Full Twitter / X profile URL.",
    "tiktok": "Optional. Full TikTok profile URL.",
    "youtube": "Optional. Full YouTube channel URL.",
    "telegram": "Optional. Full Telegram URL or @handle.",
    "discord": "Optional. Discord username.",
}


# Map of canonical field -> list of accepted headers.  Matching is
# case-insensitive, accent-insensitive and ignores leading/trailing
# whitespace, dots, dashes and underscores, so "First Name", "first-name",
# "FIRST_NAME", "first.name" and even "primer nombre" all collapse to the
# same key.
COLUMN_ALIASES: dict[str, list[str]] = {
    "first_name": [
        "first name", "firstname", "first", "given name", "given",
        "nombre", "nombres", "primer nombre", "name",
    ],
    "middle_name": [
        "middle name", "middlename", "middle",
        "segundo nombre", "segundo",
    ],
    "last_name": [
        "last name", "lastname", "last", "surname", "family name", "familyname",
        "apellido", "apellidos", "apellido paterno", "apellido materno",
    ],
    "nickname": ["nickname", "nick", "alias", "apodo", "sobrenombre"],
    "gender": ["gender", "sex", "sexo", "genero"],
    "birthday": [
        "birthday", "birth date", "birthdate", "date of birth", "dob",
        "fecha de nacimiento", "fecha nacimiento", "nacimiento",
        "cumpleanios", "cumpleanos", "cumple",
    ],
    "personal_email": [
        "personal email", "personal email address", "personal",
        "private email", "email", "e-mail", "mail",
        "email personal", "correo personal", "correo",
        "email privado", "mail personal",
    ],
    "work_email": [
        "work email", "business email", "office email", "company email",
        "work mail", "email trabajo", "correo trabajo", "correo corporativo",
        "email corporativo", "email empresa", "email oficina",
    ],
    "personal_phone": [
        "personal phone", "mobile", "cell phone", "cellphone", "cellular", "cell",
        "phone", "telephone",
        "telefono personal", "movil", "celular",
        "telefono", "tel",
    ],
    "work_phone": [
        "work phone", "office phone", "business phone", "work tel",
        "extension", "ext",
        "telefono trabajo", "tel oficina",
    ],
    "company": [
        "company", "organization", "organisation", "employer", "org",
        "empresa", "compania", "organizacion",
    ],
    "website": ["website", "web", "url", "site", "sitio web", "pagina web"],
    "address": ["address", "direccion", "domicilio"],
    "country": ["country", "pais"],
    "state": ["state", "province", "region", "estado", "provincia", "region", "departamento"],
    "city": ["city", "town", "ciudad", "localidad", "municipio"],
    "notes": ["notes", "note", "description", "comments", "notas", "observaciones", "comentarios"],
    "linkedin": ["linkedin", "linkedin url", "linkedin profile", "linkedin link"],
    "github": ["github", "github url", "github profile", "github link"],
    "facebook": ["facebook", "facebook url", "facebook profile", "facebook link", "fb"],
    "instagram": ["instagram", "instagram url", "instagram profile", "instagram link", "ig"],
    "twitter": [
        "twitter", "twitter url", "twitter profile", "twitter link",
        "x", "x url", "x profile",
    ],
    "tiktok": ["tiktok", "tiktok url", "tiktok profile", "tiktok link"],
    "youtube": ["youtube", "youtube url", "youtube profile", "youtube link", "youtube channel"],
    "telegram": ["telegram", "telegram url", "telegram profile", "telegram link", "telegram handle"],
    "discord": ["discord", "discord user", "discord username", "discord handle", "discord tag"],
}


def _fold(text: str) -> str:
    """Lowercase + strip accents + collapse whitespace/punctuation."""
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    no_accents = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"[\s\-_.]+", " ", no_accents.strip().lower()).strip()


# Pre-compute alias -> canonical field lookup once.
_ALIAS_LOOKUP: dict[str, str] = {}
for _canonical, _aliases in COLUMN_ALIASES.items():
    for _alias in _aliases:
        _ALIAS_LOOKUP[_fold(_alias)] = _canonical


def _alias_lookup() -> dict[str, str]:
    return _ALIAS_LOOKUP


def _resolve_headers(headers: Iterable[str]) -> dict[int, str]:
    """Map each column index to its canonical field name (or '' if unknown)."""
    mapping: dict[int, str] = {}
    lookup = _alias_lookup()
    for idx, header in enumerate(headers):
        if header is None:
            mapping[idx] = ""
            continue
        key = _fold(str(header))
        mapping[idx] = lookup.get(key, "")
    return mapping


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _get(row: Mapping[int, Any], header_map: Mapping[int, str], field: str, default: str = "") -> str:
    for idx, canonical in header_map.items():
        if canonical == field:
            value = row.get(idx, default)
            if value is None:
                return default
            return str(value).strip()
    return default


def _parse_date(raw: str) -> date | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw.split("T")[0]).date()
    except ValueError:
        return None


def _normalize_gender(raw: str) -> str:
    raw = (raw or "").strip().lower()
    table = {
        "male": "MALE", "m": "MALE", "masculino": "MALE", "hombre": "MALE",
        "female": "FEMALE", "f": "FEMALE", "femenino": "FEMALE", "mujer": "FEMALE",
        "other": "OTHER", "otro": "OTHER", "otros": "OTHER",
        "prefer not to say": "PREFER_NOT_TO_SAY",
        "prefer not to say.": "PREFER_NOT_TO_SAY",
        "prefer_no_say": "PREFER_NOT_TO_SAY",
        "prefiero no decir": "PREFER_NOT_TO_SAY",
        "no especifica": "PREFER_NOT_TO_SAY",
    }
    return table.get(raw, "")


def _social_from_url(url: str) -> str | None:
    if not url:
        return None
    low = url.lower()
    rules = [
        ("linkedin.com", "linkedin"),
        ("github.com", "github"),
        ("facebook.com", "facebook"),
        ("fb.com", "facebook"),
        ("instagram.com", "instagram"),
        ("twitter.com", "twitter"),
        ("x.com", "twitter"),
        ("tiktok.com", "tiktok"),
        ("youtube.com", "youtube"),
        ("youtu.be", "youtube"),
        ("t.me", "telegram"),
        ("telegram.me", "telegram"),
    ]
    for needle, field in rules:
        if needle in low:
            return field
    return None


# ---------------------------------------------------------------------------
# Row -> Contact conversion (shared by CSV / XLSX / VCF / JSON)
# ---------------------------------------------------------------------------
@dataclass
class ParsedContact:
    data: dict[str, Any]
    custom_socials: list[dict[str, str]]


def _build_contact_data(
    *,
    first_name: str,
    middle_name: str = "",
    last_name: str = "",
    nickname: str = "",
    gender: str = "",
    birthday: str = "",
    personal_email: str = "",
    work_email: str = "",
    personal_phone: str = "",
    work_phone: str = "",
    company: str = "",
    website: str = "",
    address: str = "",
    country: str = "",
    state: str = "",
    city: str = "",
    notes: str = "",
    linkedin: str = "",
    github: str = "",
    facebook: str = "",
    instagram: str = "",
    twitter: str = "",
    tiktok: str = "",
    youtube: str = "",
    telegram: str = "",
    discord: str = "",
    extra_socials: list[tuple[str, str]] | None = None,
) -> ParsedContact | None:
    first_name = (first_name or "").strip()
    if not first_name:
        return None

    socials: list[tuple[str, str]] = [
        (field, value.strip())
        for field, value in (
            ("linkedin", linkedin),
            ("github", github),
            ("facebook", facebook),
            ("instagram", instagram),
            ("twitter", twitter),
            ("tiktok", tiktok),
            ("youtube", youtube),
            ("telegram", telegram),
        )
        if value and value.strip()
    ]
    if discord and discord.strip():
        socials.append(("discord", discord.strip()))
    if extra_socials:
        socials.extend(extra_socials)

    standard_social_fields = {
        "linkedin", "github", "facebook", "instagram",
        "twitter", "tiktok", "youtube", "telegram", "discord",
    }
    standard: dict[str, str] = {}
    custom: list[dict[str, str]] = []
    seen_standard: set[str] = set()
    for field, value in socials:
        if field == "discord":
            standard["discord"] = value
            continue
        if field in standard_social_fields:
            if field in seen_standard:
                continue
            seen_standard.add(field)
            standard[field] = value
        else:
            if field and value:
                custom.append({"name": field[:100], "url": value[:500]})

    data = {
        "first_name": first_name,
        "middle_name": middle_name.strip(),
        "last_name": last_name.strip(),
        "nickname": nickname.strip(),
        "gender": _normalize_gender(gender),
        "birthday": _parse_date(birthday),
        "personal_email": personal_email.strip(),
        "work_email": work_email.strip(),
        "personal_phone": personal_phone.strip(),
        "work_phone": work_phone.strip(),
        "company_name": company.strip(),
        "website": website.strip(),
        "address": address.strip(),
        "country_name": country.strip(),
        "state_name": state.strip(),
        "city_name": city.strip(),
        "notes": notes.strip(),
        "linkedin": standard.get("linkedin", ""),
        "github": standard.get("github", ""),
        "facebook": standard.get("facebook", ""),
        "instagram": standard.get("instagram", ""),
        "twitter": standard.get("twitter", ""),
        "tiktok": standard.get("tiktok", ""),
        "youtube": standard.get("youtube", ""),
        "telegram": standard.get("telegram", ""),
        "discord": standard.get("discord", ""),
    }
    return ParsedContact(data=data, custom_socials=custom)


def _persist_contact(parsed: ParsedContact, *, workspace, user) -> Contact:
    d = parsed.data
    company_obj = None
    if d["company_name"]:
        company_obj = get_or_create_locale(
            Company,
            d["company_name"],
            parent_field="workspace",
            extra_defaults={"workspace": workspace},
            code_length=0,
        )

    country_obj = None
    state_obj = None
    city_obj = None
    if d["country_name"]:
        country_obj = get_or_create_locale(Country, d["country_name"])
        if country_obj and d["state_name"]:
            state_obj = get_or_create_locale(State, d["state_name"], parent=country_obj)
            if state_obj and d["city_name"]:
                city_obj = get_or_create_locale(
                    City,
                    d["city_name"],
                    parent=state_obj,
                    parent_field="state",
                    code_length=0,
                )

    return Contact.objects.create(
        workspace=workspace,
        created_by=user,
        first_name=d["first_name"],
        middle_name=d["middle_name"],
        last_name=d["last_name"],
        nickname=d["nickname"],
        gender=d["gender"],
        birthday=d["birthday"],
        personal_email=d["personal_email"],
        work_email=d["work_email"],
        personal_phone=d["personal_phone"],
        work_phone=d["work_phone"],
        company=company_obj,
        website=d["website"],
        address=d["address"],
        country=country_obj,
        state=state_obj,
        city=city_obj,
        notes=d["notes"],
        linkedin=d["linkedin"],
        github=d["github"],
        facebook=d["facebook"],
        instagram=d["instagram"],
        twitter=d["twitter"],
        tiktok=d["tiktok"],
        youtube=d["youtube"],
        telegram=d["telegram"],
        discord=d["discord"],
        custom_social_profiles=parsed.custom_socials,
    )


# ---------------------------------------------------------------------------
# Contact -> dict (used by all exporters)
# ---------------------------------------------------------------------------
def _contact_to_row(contact: Contact) -> dict[str, str]:
    return {
        "first_name": contact.first_name,
        "middle_name": contact.middle_name or "",
        "last_name": contact.last_name or "",
        "nickname": contact.nickname or "",
        "gender": contact.gender or "",
        "birthday": contact.birthday.isoformat() if contact.birthday else "",
        "personal_email": contact.personal_email or "",
        "work_email": contact.work_email or "",
        "personal_phone": contact.personal_phone or "",
        "work_phone": contact.work_phone or "",
        "company": contact.company.name if contact.company else "",
        "website": contact.website or "",
        "address": contact.address or "",
        "country": contact.country.name_es if contact.country else "",
        "state": contact.state.name_es if contact.state else "",
        "city": contact.city.name_es if contact.city else "",
        "notes": contact.notes or "",
        "linkedin": contact.linkedin or "",
        "github": contact.github or "",
        "facebook": contact.facebook or "",
        "instagram": contact.instagram or "",
        "twitter": contact.twitter or "",
        "tiktok": contact.tiktok or "",
        "youtube": contact.youtube or "",
        "telegram": contact.telegram or "",
        "discord": contact.discord or "",
    }


# ---------------------------------------------------------------------------
# CSV import / export
# ---------------------------------------------------------------------------
def _iter_csv_rows(file_or_text) -> Iterable[list[str]]:
    if hasattr(file_or_text, "read"):
        raw = file_or_text.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8-sig", errors="replace")
        try:
            file_or_text.seek(0)
        except Exception:  # noqa: BLE001
            pass
        text = raw
    else:
        text = file_or_text

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    for row in reader:
        yield row


def import_contacts_csv(file_obj, *, workspace, user) -> dict:
    rows = list(_iter_csv_rows(file_obj))
    if not rows:
        return {"created": 0, "skipped": 0, "errors": ["The file is empty."], "total": 0}

    header_map = _resolve_headers(rows[0])

    if not any(header_map.values()):
        return {
            "created": 0,
            "skipped": 0,
            "errors": [
                "None of the columns could be matched. "
                "Download the template to see the expected column names."
            ],
            "total": 0,
        }

    return _import_rows(rows[1:], header_map, workspace=workspace, user=user, source="CSV")


def import_contacts_xlsx(file_obj, *, workspace, user) -> dict:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(filename=file_obj, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return {
            "created": 0,
            "skipped": 0,
            "errors": [f"Could not read the Excel file: {exc}"],
            "total": 0,
        }
    sheet = wb.active

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return {"created": 0, "skipped": 0, "errors": ["The file is empty."], "total": 0}

    header_map = _resolve_headers(header)
    if not any(header_map.values()):
        return {
            "created": 0,
            "skipped": 0,
            "errors": [
                "None of the columns could be matched. "
                "Download the template to see the expected column names."
            ],
            "total": 0,
        }

    rows: list[list[Any]] = []
    for row in rows_iter:
        rows.append(list(row))
    return _import_rows(rows, header_map, workspace=workspace, user=user, source="XLSX")


def _import_rows(
    rows: Iterable[Iterable[Any]],
    header_map: dict[int, str],
    *,
    workspace,
    user,
    source: str,
) -> dict:
    created = 0
    skipped = 0
    errors: list[str] = []
    total = 0
    for line_no, row in enumerate(rows, start=2):
        if not row or all((cell is None or str(cell).strip() == "") for cell in row):
            continue
        total += 1
        try:
            row_dict: dict[int, Any] = {idx: row[idx] for idx in range(len(row))}
            parsed = _build_contact_data(
                first_name=_get(row_dict, header_map, "first_name"),
                middle_name=_get(row_dict, header_map, "middle_name"),
                last_name=_get(row_dict, header_map, "last_name"),
                nickname=_get(row_dict, header_map, "nickname"),
                gender=_get(row_dict, header_map, "gender"),
                birthday=_get(row_dict, header_map, "birthday"),
                personal_email=_get(row_dict, header_map, "personal_email"),
                work_email=_get(row_dict, header_map, "work_email"),
                personal_phone=_get(row_dict, header_map, "personal_phone"),
                work_phone=_get(row_dict, header_map, "work_phone"),
                company=_get(row_dict, header_map, "company"),
                website=_get(row_dict, header_map, "website"),
                address=_get(row_dict, header_map, "address"),
                country=_get(row_dict, header_map, "country"),
                state=_get(row_dict, header_map, "state"),
                city=_get(row_dict, header_map, "city"),
                notes=_get(row_dict, header_map, "notes"),
                linkedin=_get(row_dict, header_map, "linkedin"),
                github=_get(row_dict, header_map, "github"),
                facebook=_get(row_dict, header_map, "facebook"),
                instagram=_get(row_dict, header_map, "instagram"),
                twitter=_get(row_dict, header_map, "twitter"),
                tiktok=_get(row_dict, header_map, "tiktok"),
                youtube=_get(row_dict, header_map, "youtube"),
                telegram=_get(row_dict, header_map, "telegram"),
                discord=_get(row_dict, header_map, "discord"),
            )
            if parsed is None:
                skipped += 1
                errors.append(f"Row {line_no}: first_name is required.")
                continue
            _persist_contact(parsed, workspace=workspace, user=user)
            created += 1
        except Exception as exc:  # noqa: BLE001
            skipped += 1
            errors.append(f"Row {line_no}: {exc}")
    return {"created": created, "skipped": skipped, "errors": errors[:25], "total": total, "source": source}


# ---------------------------------------------------------------------------
# vCard 3.0 import
# ---------------------------------------------------------------------------
def _vcard_unfold(text: str) -> str:
    return re.sub(r"\r?\n[ \t]", "", text)


def _vcard_split_properties(unfolded: str) -> list[tuple[str, dict[str, str], str]]:
    out = []
    for line in unfolded.splitlines():
        if not line.strip():
            continue
        if ":" not in line:
            continue
        head, value = line.split(":", 1)
        parts = head.split(";")
        key = parts[0].upper()
        params: dict[str, str] = {}
        for part in parts[1:]:
            if "=" in part:
                pk, pv = part.split("=", 1)
                params[pk.upper()] = pv
        out.append((key, params, value))
    return out


def import_contacts_vcf(file_obj, *, workspace, user) -> dict:
    if hasattr(file_obj, "read"):
        raw = file_obj.read()
        if isinstance(raw, bytes):
            text = raw.decode("utf-8-sig", errors="replace")
        else:
            text = raw
    else:
        text = file_obj

    text = _vcard_unfold(text)
    blocks = re.split(r"(?im)^BEGIN:VCARD\s*$", text)
    cards = []
    for block in blocks:
        if "END:VCARD" not in block.upper():
            continue
        body = block.split("END:VCARD", 1)[0]
        cards.append(body)

    if not cards:
        return {"created": 0, "skipped": 0, "errors": ["No vCard (BEGIN:VCARD) blocks found."], "total": 0}

    created = 0
    skipped = 0
    errors: list[str] = []
    total = 0
    for idx, card in enumerate(cards, start=1):
        total += 1
        try:
            props = _vcard_split_properties(card)
            if not props:
                skipped += 1
                errors.append(f"Card #{idx}: empty or unreadable.")
                continue

            first_name = ""
            middle_name = ""
            last_name = ""
            nickname = ""
            gender = ""
            birthday = ""
            personal_email = ""
            work_email = ""
            personal_phone = ""
            work_phone = ""
            company = ""
            website = ""
            address = ""
            country = ""
            state = ""
            city = ""
            notes = ""
            social_urls: dict[str, str] = {}
            discord = ""
            extra_socials: list[tuple[str, str]] = []

            for key, params, value in props:
                v = value.strip()
                ptype = (params.get("TYPE") or "").upper()
                types = {t.strip().upper() for t in ptype.split(",") if t.strip()}

                if key == "N":
                    parts = v.split(";")
                    if len(parts) >= 1:
                        last_name = parts[0]
                    if len(parts) >= 2:
                        first_name = parts[1]
                    if len(parts) >= 3:
                        middle_name = parts[2]
                elif key == "FN" and not first_name:
                    parts = v.split(" ", 1)
                    first_name = parts[0]
                    if not last_name and len(parts) > 1:
                        last_name = parts[1]
                elif key == "NICKNAME":
                    nickname = v.split(",")[0]
                elif key == "GENDER":
                    gender = v
                elif key == "BDAY":
                    birthday = v[:10]
                elif key == "EMAIL":
                    if "WORK" in types:
                        work_email = v
                    else:
                        personal_email = v or personal_email
                elif key == "TEL":
                    if "WORK" in types:
                        work_phone = v
                    elif "FAX" in types:
                        pass
                    else:
                        personal_phone = v or personal_phone
                elif key == "ORG":
                    company = v.split(";")[0]
                elif key == "TITLE":
                    if not notes:
                        notes = v
                elif key == "URL":
                    hinted = None
                    for t in types:
                        t_low = t.lower()
                        if t_low in {
                            "linkedin", "github", "facebook", "instagram",
                            "twitter", "tiktok", "youtube", "telegram",
                        }:
                            hinted = t_low
                            break
                    social_field = hinted or _social_from_url(v)
                    if not website:
                        website = v
                    elif social_field:
                        social_urls.setdefault(social_field, v)
                    else:
                        extra_socials.append(("URL", v))
                elif key == "ADR":
                    parts = v.split(";")
                    adr_pieces = [p for p in (parts[2], parts[3], parts[4], parts[6]) if p]
                    if adr_pieces:
                        address = ", ".join(adr_pieces)
                    if len(parts) >= 7 and parts[6]:
                        country = parts[6]
                    if len(parts) >= 5 and parts[4]:
                        state = parts[4]
                    if len(parts) >= 4 and parts[3]:
                        city = parts[3]
                elif key == "NOTE":
                    notes = v
                elif key in ("X-SKYPE", "X-SKYPE-USERNAME"):
                    extra_socials.append(("Skype", v))
                elif key == "X-TWITTER":
                    social_urls.setdefault("twitter", v)
                elif key in ("X-INSTAGRAM", "X-INSTAGRAM-USERNAME"):
                    social_urls.setdefault("instagram", v)
                elif key == "X-LINKEDIN":
                    social_urls.setdefault("linkedin", v)
                elif key == "X-GITHUB":
                    social_urls.setdefault("github", v)
                elif key == "X-FACEBOOK":
                    social_urls.setdefault("facebook", v)
                elif key == "X-TELEGRAM":
                    social_urls.setdefault("telegram", v)
                elif key == "X-DISCORD":
                    discord = v
                elif key.startswith("X-") and v:
                    extra_socials.append((key[2:].title(), v))

            parsed = _build_contact_data(
                first_name=first_name,
                middle_name=middle_name,
                last_name=last_name,
                nickname=nickname,
                gender=gender,
                birthday=birthday,
                personal_email=personal_email,
                work_email=work_email,
                personal_phone=personal_phone,
                work_phone=work_phone,
                company=company,
                website=website,
                address=address,
                country=country,
                state=state,
                city=city,
                notes=notes,
                linkedin=social_urls.get("linkedin", ""),
                github=social_urls.get("github", ""),
                facebook=social_urls.get("facebook", ""),
                instagram=social_urls.get("instagram", ""),
                twitter=social_urls.get("twitter", ""),
                tiktok=social_urls.get("tiktok", ""),
                youtube=social_urls.get("youtube", ""),
                telegram=social_urls.get("telegram", ""),
                discord=discord,
                extra_socials=extra_socials,
            )
            if parsed is None:
                skipped += 1
                errors.append(f"Card #{idx}: FN/N missing, cannot determine a name.")
                continue
            _persist_contact(parsed, workspace=workspace, user=user)
            created += 1
        except Exception as exc:  # noqa: BLE001
            skipped += 1
            errors.append(f"Card #{idx}: {exc}")

    return {"created": created, "skipped": skipped, "errors": errors[:25], "total": total, "source": "VCF"}


# ---------------------------------------------------------------------------
# JSON import
# ---------------------------------------------------------------------------
def _coerce_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return ""
    return str(value).strip()


def _coerce_org(value: Any) -> str:
    if isinstance(value, list):
        return _coerce_str(value[0] if value else "")
    if isinstance(value, str):
        return value.strip()
    return ""


def import_contacts_json(file_obj, *, workspace, user) -> dict:
    if hasattr(file_obj, "read"):
        raw = file_obj.read()
        if isinstance(raw, bytes):
            text = raw.decode("utf-8-sig", errors="replace")
        else:
            text = raw
    else:
        text = file_obj
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        return {"created": 0, "skipped": 0, "errors": [f"Invalid JSON: {exc}"], "total": 0}

    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        return {"created": 0, "skipped": 0, "errors": ["JSON must be a list of contacts or a single contact object."], "total": 0}

    created = 0
    skipped = 0
    errors: list[str] = []
    total = 0
    for idx, item in enumerate(data, start=1):
        if not isinstance(item, dict):
            skipped += 1
            errors.append(f"Item #{idx}: not an object.")
            continue
        total += 1
        try:
            parsed = _build_contact_data(
                first_name=_coerce_str(item.get("first_name") or item.get("firstName")),
                middle_name=_coerce_str(item.get("middle_name") or item.get("middleName")),
                last_name=_coerce_str(item.get("last_name") or item.get("lastName")),
                nickname=_coerce_str(item.get("nickname")),
                gender=_coerce_str(item.get("gender")),
                birthday=_coerce_str(item.get("birthday") or item.get("bday")),
                personal_email=_coerce_str(item.get("personal_email") or item.get("personalEmail") or item.get("email")),
                work_email=_coerce_str(item.get("work_email") or item.get("workEmail")),
                personal_phone=_coerce_str(item.get("personal_phone") or item.get("personalPhone") or item.get("phone")),
                work_phone=_coerce_str(item.get("work_phone") or item.get("workPhone")),
                company=_coerce_org(item.get("company") or item.get("org")),
                website=_coerce_str(item.get("website") or item.get("url")),
                address=_coerce_str(item.get("address")),
                country=_coerce_str(item.get("country")),
                state=_coerce_str(item.get("state") or item.get("region")),
                city=_coerce_str(item.get("city")),
                notes=_coerce_str(item.get("notes") or item.get("note")),
                linkedin=_coerce_str(item.get("linkedin")),
                github=_coerce_str(item.get("github")),
                facebook=_coerce_str(item.get("facebook")),
                instagram=_coerce_str(item.get("instagram")),
                twitter=_coerce_str(item.get("twitter")),
                tiktok=_coerce_str(item.get("tiktok")),
                youtube=_coerce_str(item.get("youtube")),
                telegram=_coerce_str(item.get("telegram")),
                discord=_coerce_str(item.get("discord")),
            )
            if parsed is None:
                skipped += 1
                errors.append(f"Item #{idx}: first_name is required.")
                continue
            _persist_contact(parsed, workspace=workspace, user=user)
            created += 1
        except Exception as exc:  # noqa: BLE001
            skipped += 1
            errors.append(f"Item #{idx}: {exc}")

    return {"created": created, "skipped": skipped, "errors": errors[:25], "total": total, "source": "JSON"}


# ---------------------------------------------------------------------------
# Exporters (return HttpResponse)
# ---------------------------------------------------------------------------
def export_contacts_csv(contacts) -> HttpResponse:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in CONTACT_FIELDS])
    for c in contacts:
        row = _contact_to_row(c)
        writer.writerow([row[canonical] for canonical, _ in CONTACT_FIELDS])
    response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="contacts.csv"'
    return response


def export_contacts_xlsx(contacts) -> HttpResponse:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append([label for _, label in CONTACT_FIELDS])
    for c in contacts:
        row = _contact_to_row(c)
        ws.append([row[canonical] for canonical, _ in CONTACT_FIELDS])
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = min(max(length + 2, 12), 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="contacts.xlsx"'
    wb.save(response)
    return response


def _vcard_escape(value: str) -> str:
    return (
        (value or "")
        .replace("\\", "\\\\")
        .replace("\n", "\\n")
        .replace(",", "\\,")
        .replace(";", "\\;")
    )


def export_contacts_vcf(contacts) -> HttpResponse:
    lines: list[str] = []
    for c in contacts:
        last = c.last_name or ""
        first = c.first_name or ""
        middle = c.middle_name or ""
        lines.append("BEGIN:VCARD")
        lines.append("VERSION:3.0")
        full = " ".join(p for p in (first, middle, last) if p)
        lines.append(f"FN:{_vcard_escape(full)}")
        lines.append(
            f"N:{_vcard_escape(last)};{_vcard_escape(first)};{_vcard_escape(middle)};;"
        )
        if c.nickname:
            lines.append(f"NICKNAME:{_vcard_escape(c.nickname)}")
        if c.gender:
            lines.append(f"GENDER:{_vcard_escape(c.gender)}")
        if c.birthday:
            lines.append(f"BDAY:{c.birthday.isoformat()}")
        if c.personal_email:
            lines.append(f"EMAIL;TYPE=HOME,INTERNET:{_vcard_escape(c.personal_email)}")
        if c.work_email:
            lines.append(f"EMAIL;TYPE=WORK,INTERNET:{_vcard_escape(c.work_email)}")
        if c.personal_phone:
            lines.append(f"TEL;TYPE=CELL,VOICE:{_vcard_escape(c.personal_phone)}")
        if c.work_phone:
            lines.append(f"TEL;TYPE=WORK,VOICE:{_vcard_escape(c.work_phone)}")
        if c.company:
            lines.append(f"ORG:{_vcard_escape(c.company.name)}")
        if c.website:
            lines.append(f"URL:{_vcard_escape(c.website)}")
        if c.address:
            lines.append(f"ADR;TYPE=HOME:;;{_vcard_escape(c.address)};;;;")
        if c.notes:
            lines.append(f"NOTE:{_vcard_escape(c.notes)}")
        for field, url in (
            ("linkedin", c.linkedin),
            ("github", c.github),
            ("facebook", c.facebook),
            ("instagram", c.instagram),
            ("twitter", c.twitter),
            ("tiktok", c.tiktok),
            ("youtube", c.youtube),
            ("telegram", c.telegram),
        ):
            if url:
                lines.append(f"URL;TYPE={field.upper()}:{_vcard_escape(url)}")
        if c.discord:
            lines.append(f"X-DISCORD:{_vcard_escape(c.discord)}")
        for profile in (c.custom_social_profiles or []):
            if isinstance(profile, dict) and profile.get("name") and profile.get("url"):
                lines.append(
                    f"X-{_vcard_escape(str(profile['name']).upper())}:{_vcard_escape(str(profile['url']))}"
                )
        lines.append("END:VCARD")
    response = HttpResponse("\r\n".join(lines) + "\r\n", content_type="text/vcard; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="contacts.vcf"'
    return response


def export_contacts_json(contacts) -> HttpResponse:
    payload = []
    for c in contacts:
        row = _contact_to_row(c)
        row["social_profiles"] = c.custom_social_profiles or []
        payload.append(row)
    response = HttpResponse(
        json.dumps(payload, indent=2, ensure_ascii=False),
        content_type="application/json; charset=utf-8",
    )
    response["Content-Disposition"] = 'attachment; filename="contacts.json"'
    return response


# ---------------------------------------------------------------------------
# Template generators (CSV / XLSX / VCF)
# ---------------------------------------------------------------------------
_SAMPLE_ROW: dict[str, str] = {
    "first_name": "Jane",
    "middle_name": "",
    "last_name": "Doe",
    "nickname": "Jan",
    "gender": "FEMALE",
    "birthday": "1990-05-15",
    "personal_email": "[email protected]",
    "work_email": "[email protected]",
    "personal_phone": "+1 555 0100",
    "work_phone": "+1 555 0101",
    "company": "Acme Inc.",
    "website": "https://janedoe.com",
    "address": "123 Main St, Springfield",
    "country": "United States",
    "state": "Illinois",
    "city": "Springfield",
    "notes": "Met at the 2026 conference.",
    "linkedin": "https://linkedin.com/in/janedoe",
    "github": "https://github.com/janedoe",
    "facebook": "",
    "instagram": "https://instagram.com/janedoe",
    "twitter": "",
    "tiktok": "",
    "youtube": "",
    "telegram": "@janedoe",
    "discord": "janedoe#0001",
}


def template_csv() -> HttpResponse:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in CONTACT_FIELDS])
    writer.writerow([_SAMPLE_ROW.get(canonical, "") for canonical, _ in CONTACT_FIELDS])
    response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="contacts_template.csv"'
    return response


def template_xlsx() -> HttpResponse:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    headers = [label for _, label in CONTACT_FIELDS]
    ws.append(headers)
    ws.append([_SAMPLE_ROW.get(canonical, "") for canonical, _ in CONTACT_FIELDS])
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = min(max(length + 2, 12), 50)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="contacts_template.xlsx"'
    wb.save(response)
    return response


def template_vcf() -> HttpResponse:
    lines = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        f"FN:{_vcard_escape(_SAMPLE_ROW['first_name'] + ' ' + _SAMPLE_ROW['last_name'])}",
        f"N:{_vcard_escape(_SAMPLE_ROW['last_name'])};{_vcard_escape(_SAMPLE_ROW['first_name'])};;",
    ]
    if _SAMPLE_ROW.get("nickname"):
        lines.append(f"NICKNAME:{_vcard_escape(_SAMPLE_ROW['nickname'])}")
    if _SAMPLE_ROW.get("birthday"):
        lines.append(f"BDAY:{_SAMPLE_ROW['birthday']}")
    if _SAMPLE_ROW.get("personal_email"):
        lines.append(f"EMAIL;TYPE=HOME,INTERNET:{_vcard_escape(_SAMPLE_ROW['personal_email'])}")
    if _SAMPLE_ROW.get("work_email"):
        lines.append(f"EMAIL;TYPE=WORK,INTERNET:{_vcard_escape(_SAMPLE_ROW['work_email'])}")
    if _SAMPLE_ROW.get("personal_phone"):
        lines.append(f"TEL;TYPE=CELL,VOICE:{_vcard_escape(_SAMPLE_ROW['personal_phone'])}")
    if _SAMPLE_ROW.get("company"):
        lines.append(f"ORG:{_vcard_escape(_SAMPLE_ROW['company'])}")
    if _SAMPLE_ROW.get("website"):
        lines.append(f"URL:{_vcard_escape(_SAMPLE_ROW['website'])}")
    if _SAMPLE_ROW.get("linkedin"):
        lines.append(f"URL;TYPE=LINKEDIN:{_vcard_escape(_SAMPLE_ROW['linkedin'])}")
    if _SAMPLE_ROW.get("notes"):
        lines.append(f"NOTE:{_vcard_escape(_SAMPLE_ROW['notes'])}")
    lines.append("END:VCARD")
    response = HttpResponse("\r\n".join(lines) + "\r\n", content_type="text/vcard; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="contacts_template.vcf"'
    return response


# ---------------------------------------------------------------------------
# Dispatcher helpers used by the view
# ---------------------------------------------------------------------------
IMPORTERS = {
    "csv": import_contacts_csv,
    "xlsx": import_contacts_xlsx,
    "vcf": import_contacts_vcf,
    "json": import_contacts_json,
}

EXPORTERS = {
    "csv": export_contacts_csv,
    "xlsx": export_contacts_xlsx,
    "vcf": export_contacts_vcf,
    "json": export_contacts_json,
}

TEMPLATES = {
    "csv": template_csv,
    "xlsx": template_xlsx,
    "vcf": template_vcf,
}
